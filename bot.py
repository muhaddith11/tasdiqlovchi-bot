import logging
import requests
import openpyxl
import io
import json
import os
import datetime
import time
import pytz
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand
from telegram.ext import Application, MessageHandler, CallbackQueryHandler, CommandHandler, filters, ContextTypes

BOT_TOKEN = "8847266024:AAGA00Bqrw3ekbo5TCSmusK3Yd0FU2exTsM"
ADMIN_ID = 807823872
ONEDRIVE_URL = "https://1drv.ms/x/c/0434e9c0edef097b/IQASHiM8IYUQSZNJNl0nojFBAcv7R4dXvdm4vdX1NQN-AJw?e=38Ffl7&download=1"
TASHKENT_TZ = pytz.timezone('Asia/Tashkent')
DATA_FILE = "data.json"

logging.basicConfig(level=logging.INFO)
pending = {}
ALLOWED_COMMANDS = ['/kirim', '/chiqim']


# ─── Ma'lumotlarni saqlash ────────────────────────────────────────────────────

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE) as f:
            return json.load(f)
    return {}

def save_data(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f)

def get_group_chat_id():
    return load_data().get('group_chat_id')

def set_group_chat_id(chat_id):
    data = load_data()
    data['group_chat_id'] = chat_id
    save_data(data)


# ─── Excel funksiyalari ───────────────────────────────────────────────────────

def excel_yukla():
    url = ONEDRIVE_URL + f"&nocache={int(time.time())}"
    r = requests.get(url, timeout=30, headers={"Cache-Control": "no-cache"})
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)


def get_kunlik_tushum():
    try:
        wb = excel_yukla()
        today = datetime.datetime.now(TASHKENT_TZ).date()
        sheets = ['Салом сити-1', 'Салом сити-2', 'МЖК-1', 'МЖК-2']
        results = {}

        for sheet_name in sheets:
            ws = wb[sheet_name]
            total = 0
            for row in ws.iter_rows(values_only=True):
                date_cell = row[9] if len(row) > 9 else None    # J ustun - sana
                amount_cell = row[10] if len(row) > 10 else None # K ustun - summa
                if isinstance(date_cell, datetime.datetime) and date_cell.date() == today:
                    total += (amount_cell or 0)
            results[sheet_name] = total

        jami = sum(results.values())
        today_str = today.strftime("%d.%m.%Y")
        text = (
            f"📊 Кунлик тушум\n"
            f"📅 Сана: {today_str}\n"
            f"{'─' * 30}\n"
        )
        for name, val in results.items():
            text += f"▪️ {name:<18} {val:>10,.0f}\n"
        text += f"{'─' * 30}\n💰 Жами:               {jami:>10,.0f}"
        return text

    except Exception as e:
        logging.error(f"Kunlik tushum xatosi: {e}")
        return "❌ Ma'lumot olishda xatolik yuz berdi."


def get_bugungi_tulumlar():
    """Har bir varaq uchun alohida xabar ro'yxatini qaytaradi"""
    try:
        wb = excel_yukla()
        today = datetime.datetime.now(TASHKENT_TZ).date()
        sheets = ['Салом сити-1', 'Салом сити-2', 'МЖК-1', 'МЖК-2']
        messages = []  # Har bir varaq uchun alohida xabar

        for sheet_name in sheets:
            ws = wb[sheet_name]
            current_apt = None
            sheet_payments = []

            payment_count = 0  # joriy kvartira uchun to'lovlar soni

            for row in ws.iter_rows(min_row=7, values_only=True):
                if len(row) < 11:
                    continue
                fio = row[5]
                if fio and isinstance(fio, str) and fio.strip() and fio.strip() != 'фио':
                    foiz_val = row[12] if len(row) > 12 else None
                    current_apt = {
                        'fio': fio.strip(),
                        'dom': row[3],
                        'etaj': row[4],
                        'tulangan': row[10],
                        'qarz': row[11],
                        'foiz': foiz_val
                    }
                    payment_count = 0  # yangi kvartira - hisobni noldan boshla

                if current_apt:
                    date_val = row[9]
                    amount = row[10]
                    if isinstance(date_val, datetime.datetime) and amount:
                        payment_count += 1  # har bir to'lov sanab boriladi
                        if date_val.date() == today:
                            sheet_payments.append({
                                **current_apt,
                                'berdi': amount,
                                'toliq_son': payment_count  # nechanchi to'lov
                            })

            if sheet_payments:
                text = f"🏢 {sheet_name} — {today.strftime('%d.%m.%Y')}\n{'─' * 30}\n\n"
                for i, p in enumerate(sheet_payments, 1):
                    foiz = p['foiz']
                    foiz_str = f"{foiz * 100:.0f}%" if isinstance(foiz, float) else "—"
                    text += (
                        f"{i}. 👤 {p['fio']}\n"
                        f"   🏠 {p['dom']}-дом, {p['etaj']}-этаж\n"
                        f"   🔢 {p['toliq_son']}-chi to'lov\n"
                        f"   💵 Bugun berdi:   ${p['berdi']:>10,.0f}\n"
                        f"   ✅ Jami to'lagan: ${p['tulangan']:>10,.0f}\n"
                        f"   ❌ Qolgan qarz:   ${p['qarz']:>10,.0f}\n"
                        f"   📊 Foizda: {foiz_str}\n\n"
                    )
                text += f"{'─' * 30}\nJami: {len(sheet_payments)} ta to'lov"
                messages.append(text)

        if not messages:
            return ["📭 Bugun hech qanday to'lov kiritilmagan"]
        return messages

    except Exception as e:
        logging.error(f"Bugungi to'lov xatosi: {e}")
        return ["❌ Ma'lumot olishda xatolik yuz berdi."]


# ─── Xabar moderatsiyasi ──────────────────────────────────────────────────────

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message
    if not message or message.chat.type not in ['group', 'supergroup']:
        return

    set_group_chat_id(message.chat_id)

    text = message.text or message.caption or ""
    if not any(text.startswith(cmd) for cmd in ALLOWED_COMMANDS):
        return

    sender = message.from_user
    sender_name = sender.full_name
    username = f"@{sender.username}" if sender.username else "username yo'q"
    chat_id = message.chat_id
    key = f"{chat_id}_{message.message_id}"

    pending[key] = {
        'sender_name': sender_name,
        'username': username,
        'chat_id': chat_id,
        'message': message
    }

    keyboard = [[
        InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_{key}"),
        InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_{key}")
    ]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    footer = f"\n\n👤 Yuboruvchi: {sender_name}\n📱 Username: {username}"

    if message.text:
        await context.bot.send_message(chat_id=ADMIN_ID, text=message.text + footer, reply_markup=reply_markup)
    elif message.photo:
        await context.bot.send_photo(chat_id=ADMIN_ID, photo=message.photo[-1].file_id, caption=(message.caption or "") + footer, reply_markup=reply_markup)
    elif message.video:
        await context.bot.send_video(chat_id=ADMIN_ID, video=message.video.file_id, caption=(message.caption or "") + footer, reply_markup=reply_markup)
    elif message.document:
        await context.bot.send_document(chat_id=ADMIN_ID, document=message.document.file_id, caption=(message.caption or "") + footer, reply_markup=reply_markup)


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    action, key = query.data.split('_', 1)

    if key not in pending:
        await query.edit_message_text("Bu xabar allaqachon ko'rib chiqilgan.")
        return

    info = pending.pop(key)
    message = info['message']
    sender_name = info['sender_name']
    username = info['username']
    chat_id = info['chat_id']

    status = "✅ Boshliq tasdiqladi" if action == "approve" else "❌ Boshliq rad etdi"
    footer = f"\n\n👤 Yuboruvchi: {sender_name}\n📱 Username: {username}\n\n{status}"

    if message.text:
        await context.bot.send_message(chat_id=chat_id, text=message.text + footer)
    elif message.photo:
        await context.bot.send_photo(chat_id=chat_id, photo=message.photo[-1].file_id, caption=(message.caption or "") + footer)
    elif message.video:
        await context.bot.send_video(chat_id=chat_id, video=message.video.file_id, caption=(message.caption or "") + footer)
    elif message.document:
        await context.bot.send_document(chat_id=chat_id, document=message.document.file_id, caption=(message.caption or "") + footer)

    original_text = query.message.text or query.message.caption or ""
    new_text = original_text + f"\n\n{status}"
    if query.message.text:
        await query.edit_message_text(text=new_text)
    elif query.message.caption is not None:
        await query.edit_message_caption(caption=new_text)


# ─── Hisobot ─────────────────────────────────────────────────────────────────

async def chatid_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"Chat ID: `{update.message.chat_id}`", parse_mode='Markdown')


async def hisobot_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = await update.message.reply_text("⏳ Yuklanmoqda...")
    await msg.edit_text(get_kunlik_tushum())
    for m in get_bugungi_tulumlar():
        await update.message.reply_text(m)


async def send_daily_report(context: ContextTypes.DEFAULT_TYPE):
    # 1. Kunlik tushum jadvali
    await context.bot.send_message(chat_id=ADMIN_ID, text=get_kunlik_tushum())
    # 2. Bugungi to'lovlar - har bir varaq alohida xabar
    for msg in get_bugungi_tulumlar():
        await context.bot.send_message(chat_id=ADMIN_ID, text=msg)


# ─── Asosiy ──────────────────────────────────────────────────────────────────

async def post_init(app):
    await app.bot.set_my_commands([
        BotCommand("hisobot", "Kunlik hisobot"),
    ])
    # Har kuni soat 21:00 Toshkent = 16:00 UTC
    app.job_queue.run_daily(
        send_daily_report,
        time=datetime.time(hour=16, minute=0, tzinfo=pytz.utc)
    )


def main():
    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()

    app.add_handler(MessageHandler(filters.ChatType.GROUPS & ~filters.COMMAND, handle_message))
    app.add_handler(CommandHandler(["kirim", "chiqim"], handle_message))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(CommandHandler("hisobot", hisobot_command))
    app.add_handler(CommandHandler("chatid", chatid_command))

    app.run_polling()


if __name__ == '__main__':
    main()
