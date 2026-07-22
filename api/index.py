from flask import Flask, request, jsonify
import requests as req_lib
import json
import os
import psycopg2
import html as html_mod
import openpyxl
import io
import datetime
import time
import pytz

app = Flask(__name__)

BOT_TOKEN   = os.environ.get('BOT_TOKEN', '')
ADMIN_ID    = int(os.environ.get('ADMIN_ID', '807823872'))
GROUP_ID    = int(os.environ.get('GROUP_ID', '-1002449545348'))
DATABASE_URL = os.environ.get('DATABASE_URL', '')
ONEDRIVE_URL = os.environ.get('ONEDRIVE_URL',
    'https://1drv.ms/x/c/0434e9c0edef097b/IQASHiM8IYUQSZNJNl0nojFBAcv7R4dXvdm4vdX1NQN-AJw?e=38Ffl7&download=1')
TASHKENT_TZ      = pytz.timezone('Asia/Tashkent')
ALLOWED_COMMANDS = ['/kirim', '/chiqim']


# ─── Telegram API (raw HTTP, no async) ───────────────────────────────────────

def tg(method, **kwargs):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/{method}"
    r = req_lib.post(url, json=kwargs, timeout=10)
    return r.json()

def send_message(chat_id, text, parse_mode=None, reply_markup=None):
    p = {'chat_id': chat_id, 'text': text}
    if parse_mode:   p['parse_mode'] = parse_mode
    if reply_markup: p['reply_markup'] = reply_markup
    return tg('sendMessage', **p)

def send_photo(chat_id, photo, caption='', reply_markup=None):
    p = {'chat_id': chat_id, 'photo': photo, 'caption': caption}
    if reply_markup: p['reply_markup'] = reply_markup
    return tg('sendPhoto', **p)

def send_video(chat_id, video, caption='', reply_markup=None):
    p = {'chat_id': chat_id, 'video': video, 'caption': caption}
    if reply_markup: p['reply_markup'] = reply_markup
    return tg('sendVideo', **p)

def send_document(chat_id, document, caption='', reply_markup=None):
    p = {'chat_id': chat_id, 'document': document, 'caption': caption}
    if reply_markup: p['reply_markup'] = reply_markup
    return tg('sendDocument', **p)

def answer_callback(callback_id):
    tg('answerCallbackQuery', callback_query_id=callback_id)

def edit_message_text(chat_id, message_id, text):
    try:
        tg('editMessageText', chat_id=chat_id, message_id=message_id, text=text)
    except Exception:
        pass

def edit_message_caption(chat_id, message_id, caption):
    try:
        tg('editMessageCaption', chat_id=chat_id, message_id=message_id, caption=caption)
    except Exception:
        pass


# ─── Database ─────────────────────────────────────────────────────────────────

def get_db():
    return psycopg2.connect(DATABASE_URL)

def ensure_table():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS pending (
            key TEXT PRIMARY KEY,
            sender_name TEXT, username TEXT, chat_id BIGINT,
            msg_type TEXT, msg_text TEXT, file_id TEXT, caption TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        )''')
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"DB init error: {e}")

def db_save(key, sender_name, username, chat_id, msg_type, msg_text, file_id, caption):
    conn = get_db(); cur = conn.cursor()
    cur.execute('''INSERT INTO pending
        (key,sender_name,username,chat_id,msg_type,msg_text,file_id,caption)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (key) DO NOTHING''',
        (key, sender_name, username, chat_id, msg_type, msg_text, file_id, caption))
    conn.commit(); cur.close(); conn.close()

def db_get(key):
    conn = get_db(); cur = conn.cursor()
    cur.execute('SELECT sender_name,username,chat_id,msg_type,msg_text,file_id,caption FROM pending WHERE key=%s', (key,))
    row = cur.fetchone(); cur.close(); conn.close()
    return row

def db_delete(key):
    conn = get_db(); cur = conn.cursor()
    cur.execute('DELETE FROM pending WHERE key=%s', (key,))
    conn.commit(); cur.close(); conn.close()


# ─── Excel ────────────────────────────────────────────────────────────────────

def excel_yukla():
    url = ONEDRIVE_URL + f"&nocache={int(time.time())}"
    r = req_lib.get(url, timeout=30, headers={"Cache-Control": "no-cache"})
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)

def get_kunlik_tushum():
    try:
        wb = excel_yukla()
        today = datetime.datetime.now(TASHKENT_TZ).date()
        sheets = ['Салом сити-1', 'Салом сити-2', 'МЖК-1', 'МЖК-2']
        results = {}
        for sn in sheets:
            ws = wb[sn]; total = 0
            for row in ws.iter_rows(values_only=True):
                dc = row[9] if len(row) > 9 else None
                ac = row[10] if len(row) > 10 else None
                if isinstance(dc, datetime.datetime) and dc.date() == today:
                    total += (ac or 0)
            results[sn] = total
        jami = sum(results.values())
        text = f"📊 Кунлик тушум\n📅 Сана: {today.strftime('%d.%m.%Y')}\n{'─'*30}\n"
        for name, val in results.items():
            text += f"▪️ {name:<18} {val:>10,.0f}\n"
        text += f"{'─'*30}\n💰 Жами:               {jami:>10,.0f}"
        return text
    except Exception as e:
        return f"❌ Xatolik: {e}"

def get_bugungi_tulumlar():
    try:
        wb = excel_yukla()
        today = datetime.datetime.now(TASHKENT_TZ).date()
        sheets = ['Салом сити-1', 'Салом сити-2', 'МЖК-1', 'МЖК-2']
        messages = []
        for sn in sheets:
            ws = wb[sn]
            cur_apt = None; payments = []; pay_count = 0; last_date = None
            for row in ws.iter_rows(min_row=7, values_only=True):
                if len(row) < 11: continue
                fio = row[5]
                if fio and isinstance(fio, str) and fio.strip() and fio.strip() != 'фио':
                    cur_apt = {'fio': fio.strip(), 'kv': row[2], 'dom': row[3],
                               'etaj': row[4], 'tulangan': row[10], 'qarz': row[11],
                               'foiz': row[12] if len(row) > 12 else None}
                    pay_count = 0; last_date = None
                if cur_apt:
                    dv = row[9]; am = row[10]
                    if isinstance(dv, datetime.datetime) and am:
                        pay_count += 1
                        if dv.date() == today:
                            payments.append({**cur_apt, 'berdi': am,
                                             'toliq_son': pay_count, 'oldingi': last_date})
                        else:
                            last_date = dv.date()
            if payments:
                text = f"🏢 {sn} — {today.strftime('%d.%m.%Y')}\n{'─'*30}\n\n"
                for i, p in enumerate(payments, 1):
                    foiz = p['foiz']
                    foiz_str = f"{foiz*100:.0f}%" if isinstance(foiz, float) else "—"
                    old_str = p['oldingi'].strftime("%d.%m.%Y") if p['oldingi'] else "birinchi to'lov"
                    text += (
                        f"{i}. 👤 {html_mod.escape(str(p['fio']))}\n"
                        f"   🏠 {html_mod.escape(str(p['dom']))}-дом, "
                        f"{html_mod.escape(str(p['etaj']))}-этаж, "
                        f"{html_mod.escape(str(p['kv']))}-кв\n"
                        f"   🔢 {p['toliq_son']}-chi to'lov\n"
                        f"   📅 Oldingi to'lov: {old_str}\n"
                        f"   💵 <b>Bugun berdi:   ${p['berdi']:,.0f}</b>\n"
                        f"   ✅ Jami to'lagan: ${p['tulangan']:,.0f}\n"
                        f"   ❌ Qolgan qarz:   ${p['qarz']:,.0f}\n"
                        f"   📊 Foizda: {foiz_str}\n\n"
                    )
                text += f"{'─'*30}\nJami: {len(payments)} ta to'lov"
                messages.append(text)
        return messages if messages else ["📭 Bugun hech qanday to'lov kiritilmagan"]
    except Exception as e:
        return [f"❌ Xatolik: {e}"]


# ─── Update handlers ──────────────────────────────────────────────────────────

def handle_message(msg):
    chat_type = msg.get('chat', {}).get('type', '')
    text = msg.get('text') or msg.get('caption') or ''
    chat_id = msg['chat']['id']
    msg_id = msg['message_id']

    # /hisobot — guruh yoki shaxsiy
    if text.startswith('/hisobot'):
        for m in get_bugungi_tulumlar():
            send_message(chat_id, m, parse_mode='HTML')
        send_message(chat_id, get_kunlik_tushum())
        return

    # /chatid
    if text.startswith('/chatid'):
        send_message(chat_id, f"Chat ID: `{chat_id}`", parse_mode='Markdown')
        return

    # /kirim yoki /chiqim — faqat guruhda
    if chat_type not in ('group', 'supergroup'):
        return
    if not any(text.startswith(cmd) for cmd in ALLOWED_COMMANDS):
        return

    sender = msg.get('from', {})
    sender_name = f"{sender.get('first_name','')} {sender.get('last_name','')}" .strip()
    username = f"@{sender['username']}" if sender.get('username') else "username yo'q"
    key = f"{chat_id}_{msg_id}"

    ensure_table()

    photo = msg.get('photo')
    video = msg.get('video')
    document = msg.get('document')

    if msg.get('text'):
        msg_type, msg_text, file_id, cap = 'text', msg['text'], None, None
    elif photo:
        msg_type, msg_text, file_id, cap = 'photo', None, photo[-1]['file_id'], msg.get('caption','')
    elif video:
        msg_type, msg_text, file_id, cap = 'video', None, video['file_id'], msg.get('caption','')
    elif document:
        msg_type, msg_text, file_id, cap = 'document', None, document['file_id'], msg.get('caption','')
    else:
        return

    db_save(key, sender_name, username, chat_id, msg_type, msg_text, file_id, cap)

    keyboard = {'inline_keyboard': [[
        {'text': '✅ Tasdiqlash', 'callback_data': f'approve_{key}'},
        {'text': '❌ Rad etish',  'callback_data': f'reject_{key}'}
    ]]}
    footer = f"\n\n👤 Yuboruvchi: {sender_name}\n📱 Username: {username}"

    if msg_type == 'text':
        send_message(ADMIN_ID, msg_text + footer, reply_markup=keyboard)
    elif msg_type == 'photo':
        send_photo(ADMIN_ID, file_id, caption=(cap or '')+footer, reply_markup=keyboard)
    elif msg_type == 'video':
        send_video(ADMIN_ID, file_id, caption=(cap or '')+footer, reply_markup=keyboard)
    elif msg_type == 'document':
        send_document(ADMIN_ID, file_id, caption=(cap or '')+footer, reply_markup=keyboard)


def handle_callback(cq):
    answer_callback(cq['id'])
    data = cq.get('data', '')
    parts = data.split('_', 1)
    if len(parts) != 2:
        return
    action, key = parts

    row = db_get(key)
    if not row:
        orig_msg = cq.get('message', {})
        old_text = orig_msg.get('text') or orig_msg.get('caption') or ''
        if orig_msg.get('text'):
            edit_message_text(orig_msg['chat']['id'], orig_msg['message_id'],
                              old_text + "\n\n(allaqachon ko'rib chiqilgan)")
        return

    sender_name, username, chat_id, msg_type, msg_text, file_id, caption = row
    db_delete(key)

    status = "✅ Boshliq tasdiqladi" if action == "approve" else "❌ Boshliq rad etdi"
    footer = f"\n\n👤 Yuboruvchi: {sender_name}\n📱 Username: {username}\n\n{status}"

    if msg_type == 'text':
        send_message(chat_id, msg_text + footer)
    elif msg_type == 'photo':
        send_photo(chat_id, file_id, caption=(caption or '')+footer)
    elif msg_type == 'video':
        send_video(chat_id, file_id, caption=(caption or '')+footer)
    elif msg_type == 'document':
        send_document(chat_id, file_id, caption=(caption or '')+footer)

    orig_msg = cq.get('message', {})
    old_text = orig_msg.get('text') or orig_msg.get('caption') or ''
    new_text = old_text + f"\n\n{status}"
    if orig_msg.get('text'):
        edit_message_text(orig_msg['chat']['id'], orig_msg['message_id'], new_text)
    elif orig_msg.get('caption') is not None:
        edit_message_caption(orig_msg['chat']['id'], orig_msg['message_id'], new_text)


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/api/webhook', methods=['POST'])
def webhook():
    try:
        update = request.get_json(force=True, silent=True) or {}
        if 'message' in update:
            handle_message(update['message'])
        elif 'callback_query' in update:
            handle_callback(update['callback_query'])
    except Exception as e:
        print(f"Webhook error: {e}")
    return jsonify({"ok": True})


@app.route('/api/cron', methods=['GET'])
def cron():
    try:
        tulumlar = get_bugungi_tulumlar()
        tushum = get_kunlik_tushum()
        for chat_id in [ADMIN_ID, GROUP_ID]:
            for m in tulumlar:
                send_message(chat_id, m, parse_mode='HTML')
            send_message(chat_id, tushum)
    except Exception as e:
        print(f"Cron error: {e}")
    return jsonify({"ok": True})


@app.route('/', methods=['GET'])
def index():
    return jsonify({"status": "running"})
