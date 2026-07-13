import requests
import openpyxl
import io
import datetime
import time
import pytz
import html as html_mod
import os

ONEDRIVE_URL = os.environ.get('ONEDRIVE_URL', 'https://1drv.ms/x/c/0434e9c0edef097b/IQASHiM8IYUQSZNJNl0nojFBAcv7R4dXvdm4vdX1NQN-AJw?e=38Ffl7&download=1')
TASHKENT_TZ = pytz.timezone('Asia/Tashkent')


def excel_yukla():
    url = ONEDRIVE_URL + f"&nocache={int(time.time())}"
    r = requests.get(url, timeout=30, headers={"Cache-Control": "no-cache"})
    return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)


def get_kunlik_tushum():
    try:
        wb = excel_yukla()
        today = datetime.datetime.now(TASHKENT_TZ).date()
        sheets = ['Салом сити-1', 'Салом сити-2', 'МЖК-1', 'МЖК-2']
        results = {}
        for sheet_name in sheets:
            ws = wb[sheet_name]
            total = 0
            for row in ws.iter_rows(values_only=True):
                date_cell = row[9] if len(row) > 9 else None
                amount_cell = row[10] if len(row) > 10 else None
                if isinstance(date_cell, datetime.datetime) and date_cell.date() == today:
                    total += (amount_cell or 0)
            results[sheet_name] = total
        jami = sum(results.values())
        today_str = today.strftime("%d.%m.%Y")
        text = f"📊 Кунлик тушум\n📅 Сана: {today_str}\n{'─' * 30}\n"
        for name, val in results.items():
            text += f"▪️ {name:<18} {val:>10,.0f}\n"
        text += f"{'─' * 30}\n💰 Жами:               {jami:>10,.0f}"
        return text
    except Exception as e:
        return f"❌ Ma'lumot olishda xatolik: {e}"


def get_bugungi_tulumlar():
    try:
        wb = excel_yukla()
        today = datetime.datetime.now(TASHKENT_TZ).date()
        sheets = ['Салом сити-1', 'Салом сити-2', 'МЖК-1', 'МЖК-2']
        messages = []

        for sheet_name in sheets:
            ws = wb[sheet_name]
            current_apt = None
            sheet_payments = []
            payment_count = 0
            last_payment_date = None

            for row in ws.iter_rows(min_row=7, values_only=True):
                if len(row) < 11:
                    continue
                fio = row[5]
                if fio and isinstance(fio, str) and fio.strip() and fio.strip() != 'фио':
                    foiz_val = row[12] if len(row) > 12 else None
                    current_apt = {
                        'fio': fio.strip(),
                        'kv': row[2],
                        'dom': row[3],
                        'etaj': row[4],
                        'tulangan': row[10],
                        'qarz': row[11],
                        'foiz': foiz_val
                    }
                    payment_count = 0
                    last_payment_date = None

                if current_apt:
                    date_val = row[9]
                    amount = row[10]
                    if isinstance(date_val, datetime.datetime) and amount:
                        payment_count += 1
                        if date_val.date() == today:
                            sheet_payments.append({
                                **current_apt,
                                'berdi': amount,
                                'toliq_son': payment_count,
                                'oldingi_tulov': last_payment_date
                            })
                        else:
                            last_payment_date = date_val.date()

            if sheet_payments:
                text = f"🏢 {sheet_name} — {today.strftime('%d.%m.%Y')}\n{'─' * 30}\n\n"
                for i, p in enumerate(sheet_payments, 1):
                    foiz = p['foiz']
                    foiz_str = f"{foiz * 100:.0f}%" if isinstance(foiz, float) else "—"
                    oldingi = p['oldingi_tulov']
                    oldingi_str = oldingi.strftime("%d.%m.%Y") if oldingi else "birinchi to'lov"
                    fio = html_mod.escape(str(p['fio']))
                    dom = html_mod.escape(str(p['dom']))
                    etaj = html_mod.escape(str(p['etaj']))
                    kv = html_mod.escape(str(p['kv']))
                    text += (
                        f"{i}. 👤 {fio}\n"
                        f"   🏠 {dom}-дом, {etaj}-этаж, {kv}-кв\n"
                        f"   🔢 {p['toliq_son']}-chi to'lov\n"
                        f"   📅 Oldingi to'lov: {oldingi_str}\n"
                        f"   💵 <b>Bugun berdi:   ${p['berdi']:,.0f}</b>\n"
                        f"   ✅ Jami to'lagan: ${p['tulangan']:,.0f}\n"
                        f"   ❌ Qolgan qarz:   ${p['qarz']:,.0f}\n"
                        f"   📊 Foizda: {foiz_str}\n\n"
                    )
                text += f"{'─' * 30}\nJami: {len(sheet_payments)} ta to'lov"
                messages.append(text)

        if not messages:
            return ["📭 Bugun hech qanday to'lov kiritilmagan"]
        return messages
    except Exception as e:
        return [f"❌ Ma'lumot olishda xatolik: {e}"]
